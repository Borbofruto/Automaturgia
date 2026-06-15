"""
xlsx_generator.py — Geração de planilhas Excel (.xlsx) para entregas da Automaturgia.

Ideal para: parâmetros técnicos tabulados, datasets de benchmarks, comparações numéricas.
Requer: pip install openpyxl

Uso:
    # Via JSON na linha de comando:
    python xlsx_generator.py --json data.json --output parâmetros_dh.xlsx --title "DH JAKA Pro 16"

    # Via Python (import):
    from xlsx_generator import generate
    generate(
        sheets={
            "Parâmetros DH": [
                {"Joint": 1, "a (mm)": 0, "d (mm)": 118, "alpha (rad)": 0, "theta (rad)": 0},
                ...
            ]
        },
        output_path="output/dh_params.xlsx",
        title="Parâmetros DH JAKA Pro 16"
    )
"""

import sys
import os
import json
import argparse
from datetime import datetime


def generate(
    sheets: dict[str, list[dict]],
    output_path: str,
    title: str = "",
    author: str = "Automaturgia"
) -> str:
    """
    Gera planilha .xlsx com uma ou mais abas.

    Args:
        sheets:       Dict de { nome_aba: lista_de_dicts }
                      Cada dict representa uma linha, keys são cabeçalhos.
        output_path:  Caminho .xlsx de saída
        title:        Título da planilha (aparece na primeira célula da primeira aba)
        author:       Autor

    Returns:
        Caminho do arquivo .xlsx gerado
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Instale: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.properties.title = title
    wb.properties.creator = author
    wb.properties.created = datetime.now()

    # Remove aba padrão vazia
    default_sheet = wb.active
    sheet_names = list(sheets.keys())

    for sheet_idx, (sheet_name, rows) in enumerate(sheets.items()):
        if sheet_idx == 0:
            ws = default_sheet
            ws.title = sheet_name[:31]  # Excel limita nome de aba a 31 chars
        else:
            ws = wb.create_sheet(title=sheet_name[:31])

        if not rows:
            continue

        headers = list(rows[0].keys())

        # Estilos
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        alt_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        border = Border(
            left=Side(style="thin", color="BDC3C7"),
            right=Side(style="thin", color="BDC3C7"),
            top=Side(style="thin", color="BDC3C7"),
            bottom=Side(style="thin", color="BDC3C7")
        )

        start_row = 1

        # Título opcional na primeira linha da primeira aba
        if sheet_idx == 0 and title:
            ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
            ws.cell(row=2, column=1, value=f"Automaturgia · Gerado em {datetime.now().strftime('%d/%m/%Y')}")
            ws.cell(row=2, column=1).font = Font(italic=True, color="7F8C8D", size=9)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
            start_row = 4  # Dados começam na linha 4

        # Cabeçalhos
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Dados
        for row_idx, row_data in enumerate(rows, start_row + 1):
            is_alt = (row_idx - start_row) % 2 == 0
            for col_idx, header in enumerate(headers, 1):
                val = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if is_alt:
                    cell.fill = alt_fill

                # Números: alinha à direita
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        # Auto-ajusta largura das colunas
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(header))
            for row in ws.iter_rows(min_row=start_row + 1, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        # Congela cabeçalho
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return output_path


def from_flat_list(rows: list[dict], output_path: str, title: str = "", sheet_name: str = "Dados") -> str:
    """Atalho para gerar planilha com uma única aba a partir de lista plana de dicts."""
    return generate({sheet_name: rows}, output_path, title)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Automaturgia XLSX Generator")
    parser.add_argument("--json", help="Arquivo JSON com dados ({ sheet_name: [rows] } ou [rows])")
    parser.add_argument("--output", required=True, help="Arquivo .xlsx de saída")
    parser.add_argument("--title", default="", help="Título da planilha")
    parser.add_argument("--sheet", default="Dados", help="Nome da aba (se input for lista plana)")
    args = parser.parse_args()

    if args.json:
        with open(args.json, encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = json.loads(sys.stdin.read())

    # Aceita lista plana ou dict de sheets
    if isinstance(data, list):
        sheets = {args.sheet: data}
    else:
        sheets = data

    result = generate(sheets, args.output, args.title)
    print(f"XLSX gerado: {result}")
